from openpyxl import load_workbook
from pynput.keyboard import Key, Controller
import subprocess
import time
import os


def evaluate_excel_formula(file_path, excel_app='/Applications/Microsoft Excel.app'):
    # Ensure the file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")

    # Open the Excel file in the Excel app
    subprocess.Popen(['open', '-a', excel_app, file_path])
    time.sleep(5)  # Wait for Excel to open and calculate formulas

    # Use pynput to save the file
    keyboard = Controller()
    with keyboard.pressed(Key.cmd):
        keyboard.press('s')
        keyboard.release('s')
    time.sleep(2)

    # Use pynput to close Excel
    with keyboard.pressed(Key.cmd):
        keyboard.press('q')
        keyboard.release('q')
    time.sleep(5)  # Ensure the app closes properly

    # Load the workbook with evaluated formulas
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Display evaluated data for inspection
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    for row in data:
        print(row)

    wb.close()


# Example usage:
evaluate_excel_formula('./reports/WIPRO.xlsx')
